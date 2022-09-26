from tkinter.filedialog import askopenfile
from openpyxl import load_workbook, Workbook
from datetime import datetime
import getpass

day = datetime.today().strftime("%m%d20%y")
currentDateAndTime = datetime.now()
currentTime = currentDateAndTime.strftime("%H:%M:%S")
username = getpass.getuser()


def Data_One():
    global workbook
    try:
        workbook = Workbook()
        sheet = workbook.active

        sheet["A1"] = f"Report Date: {day}"
        sheet.append(["ID", "department", "first_name", "last_name",
                      "salary", "country", "gender"])

        for i in range(3, 8):
            sheet[f"B{i}"] = "Department: Customer Support"

    finally:
        workbook.save(f"C:\\Users\\{username}\\Desktop\\Data_One {day}_{username}.xlsx")


def Data_Two():
    global workbook
    try:
        workbook = Workbook()
        sheet = workbook.active

        sheet.append(["ID", "Department Score"])
    finally:
        workbook.save(f"C:\\Users\\{username}\\Desktop\\Data_Two_{day}_{username}.xlsx")


def Add_User_Column():
    file = askopenfile(mode="rb", title="Chose a file", filetyp=[("Excel File", "*.xlsx")])
    if file:
        workbook = load_workbook(file)
        sheet = workbook.active
        size = (sheet.max_row + 1)
        if sheet["G2"].value == "gender":
            cell_start = 3
            sheet["H2"] = "username"
            for i in range(cell_start, size):
                sheet[f"H{cell_start}"] = f"{username}"
                cell_start += 1
                workbook.save(f"C:\\Users\\marek\\Desktop\\Data_One {day}_{username}_Ready For Report.xlsx")
        if sheet["B1"].value == "Department Score":
            cell_start = 2
            sheet["H2"] = "username"
            for i in range(cell_start, size):
                sheet[f"H{cell_start}"] = f"{username}"
                cell_start += 1
                workbook.save(f"C:\\Users\\marek\\Desktop\\Data_Two {day}_{username}_Ready For Report.xlsx")
